import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook
import copy
from datetime import datetime
import os

# ============================================================
# NORĀDIET ŠEIT CEĻU UZ JŪSU EXCEL FAILU
# Piemērs: r"C:\Users\JusuVards\Documents\mans_fails.xlsx"
# ============================================================
EXCEL_FAILS = r"ŠEIT_IERAKSTIET_CELU_UZ_SAVU_EXCEL_FAILU"
# ============================================================

# Kategoriju konfigurācija
# Katrai kategorijai norādītas kolonnas Excel failā
KATEGORIJAS = {
    "Name1": {"kolonnas": [1, 2, 3],    "datums": 3,  "lauki": ["Cena", "Kas", "Datums"]},
    "Name2": {"kolonnas": [4, 5, 6],    "datums": 6,  "lauki": ["Cena", "Kas", "Datums"]},
    "Name3": {"kolonnas": [7, 8, 9],    "datums": 9,  "lauki": ["Cena", "Kas", "Datums"]},
    "Name4": {"kolonnas": [10, 11, 12], "datums": 12, "lauki": ["Cena", "Kas", "Datums"]},
    "Name5": {"kolonnas": [13, 14, 15], "datums": 15, "lauki": ["Cena", "Kas", "Datums"]},
}

# Pirmā datu rinda Excel failā (1 = pirmā rinda)
DATA_RINDA = 5

# Excel lapas nosaukums
LAPAS_NOSAUKUMS = "Lapa1"

def parse_date(val):
    if val is None:
        return datetime(9999, 1, 1)
    if isinstance(val, datetime):
        return val
    for fmt in ["%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"]:
        try:
            return datetime.strptime(str(val).strip(), fmt)
        except:
            pass
    return datetime(9999, 1, 1)

def saglabat_un_kartot(kategorija, vertibas):
    if not os.path.exists(EXCEL_FAILS):
        messagebox.showerror("Kļūda",
            f"Excel fails nav atrasts!\n\n"
            f"Lūdzu atveriet excel_paligs.py un\n"
            f"11. rindā ierakstiet pareizo ceļu uz savu failu.\n\n"
            f"Piemērs:\n"
            f'r"C:\\Users\\JusuVards\\Documents\\mans_fails.xlsx"')
        return False

    wb = load_workbook(EXCEL_FAILS)
    ws = wb[LAPAS_NOSAUKUMS]  # Excel lapas nosaukums
    info = KATEGORIJAS[kategorija]
    kolonnas = info["kolonnas"]
    datums_kol = info["datums"]
    date_idx = kolonnas.index(datums_kol)

    # Nolasīt visas esošās rindas
    rindas = []
    for rinda_nr in range(DATA_RINDA, ws.max_row + 1):
        rinda = []
        for kol in kolonnas:
            cell = ws.cell(row=rinda_nr, column=kol)
            rinda.append({
                "vertiba": cell.value,
                "font": copy.copy(cell.font),
                "fill": copy.copy(cell.fill),
                "border": copy.copy(cell.border),
                "alignment": copy.copy(cell.alignment),
                "number_format": cell.number_format,
            })
        rindas.append(rinda)

    # Pievienot jauno rindu
    jauna_rinda = []
    for i, kol in enumerate(kolonnas):
        jauna_rinda.append({
            "vertiba": vertibas[i],
            "font": copy.copy(ws.cell(row=DATA_RINDA, column=kol).font),
            "fill": copy.copy(ws.cell(row=DATA_RINDA, column=kol).fill),
            "border": copy.copy(ws.cell(row=DATA_RINDA, column=kol).border),
            "alignment": copy.copy(ws.cell(row=DATA_RINDA, column=kol).alignment),
            "number_format": ws.cell(row=DATA_RINDA, column=kol).number_format,
        })
    rindas.append(jauna_rinda)

    # Kārtot pēc datuma — vecākais augšā
    ar_datumu = [(parse_date(r[date_idx]["vertiba"]), r)
                 for r in rindas if r[date_idx]["vertiba"] is not None]
    tuksas = [r for r in rindas
              if r[date_idx]["vertiba"] is None and all(c["vertiba"] is None for c in r)]
    nav_datums = [r for r in rindas
                  if r[date_idx]["vertiba"] is None and not all(c["vertiba"] is None for c in r)]

    ar_datumu.sort(key=lambda x: x[0])
    sakartots = [r for _, r in ar_datumu] + nav_datums + tuksas

    # Paplašināt lapu ja vajadzīgs
    while len(sakartots) > ws.max_row - DATA_RINDA + 1:
        ws.append([""] * ws.max_column)

    # Ierakstīt atpakaļ Excel failā
    for offset, rinda in enumerate(sakartots):
        rinda_nr = DATA_RINDA + offset
        for col_i, kol in enumerate(kolonnas):
            cell = ws.cell(row=rinda_nr, column=kol)
            cell.value = rinda[col_i]["vertiba"]
            cell.font = rinda[col_i]["font"]
            cell.fill = rinda[col_i]["fill"]
            cell.border = rinda[col_i]["border"]
            cell.alignment = rinda[col_i]["alignment"]
            cell.number_format = rinda[col_i]["number_format"]

    wb.save(EXCEL_FAILS)
    return True


class ExcelPaligs:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Palīgs")
        self.root.geometry("480x520")
        self.root.resizable(False, False)
        self.root.configure(bg="#f5f5f5")
        self.lauki_entries = {}
        self.build_ui()

    def build_ui(self):
        # Virsraksts
        tk.Label(self.root, text="Excel Palīgs",
                 font=("Segoe UI", 16, "bold"),
                 bg="#f5f5f5", fg="#2c3e50").pack(pady=(18, 4))
        tk.Label(self.root,
                 text="Ievadiet datus — tie automātiski ieis pareizajā vietā",
                 font=("Segoe UI", 9), bg="#f5f5f5", fg="#7f8c8d").pack(pady=(0, 14))

        # Kategorijas izvēle
        frame = tk.Frame(self.root, bg="#f5f5f5")
        frame.pack(fill="x", padx=28, pady=(0, 8))
        tk.Label(frame, text="Kategorija:", font=("Segoe UI", 10, "bold"),
                 bg="#f5f5f5", width=12, anchor="w").pack(side="left")
        self.kategorija_var = tk.StringVar()
        self.kategorija_combo = ttk.Combobox(
            frame, textvariable=self.kategorija_var,
            values=list(KATEGORIJAS.keys()), state="readonly",
            font=("Segoe UI", 10), width=26)
        self.kategorija_combo.pack(side="left")
        self.kategorija_combo.bind("<<ComboboxSelected>>", self.atjaunot_laukus)

        # Ievades lauki
        self.lauki_frame = tk.Frame(self.root, bg="#f5f5f5")
        self.lauki_frame.pack(fill="x", padx=28, pady=4)

        # Poga
        self.poga = tk.Button(self.root, text="Pievienot ierakstu",
                              font=("Segoe UI", 11, "bold"),
                              bg="#27ae60", fg="white",
                              activebackground="#219a52",
                              relief="flat", cursor="hand2",
                              padx=20, pady=10,
                              command=self.pievienot)
        self.poga.pack(pady=16)

        # Statusa ziņojums
        self.status_var = tk.StringVar(value="")
        self.status_label = tk.Label(self.root,
                                      textvariable=self.status_var,
                                      font=("Segoe UI", 10), bg="#f5f5f5")
        self.status_label.pack()

    def atjaunot_laukus(self, event=None):
        for widget in self.lauki_frame.winfo_children():
            widget.destroy()
        self.lauki_entries = {}

        kategorija = self.kategorija_var.get()
        if not kategorija:
            return

        lauki = KATEGORIJAS[kategorija]["lauki"]
        for lauks in lauki:
            rinda = tk.Frame(self.lauki_frame, bg="#f5f5f5")
            rinda.pack(fill="x", pady=4)
            tk.Label(rinda, text=f"{lauks}:",
                     font=("Segoe UI", 10),
                     bg="#f5f5f5", width=12, anchor="w").pack(side="left")
            entry = tk.Entry(rinda, font=("Segoe UI", 10), width=22)
            if lauks == "Datums":
                entry.insert(0, datetime.today().strftime("%d.%m.%Y"))
            entry.pack(side="left")
            self.lauki_entries[lauks] = entry

        self.status_var.set("")

    def pievienot(self):
        kategorija = self.kategorija_var.get()
        if not kategorija:
            messagebox.showwarning("Uzmanību", "Lūdzu izvēlieties kategoriju!")
            return

        lauki = KATEGORIJAS[kategorija]["lauki"]
        vertibas = []

        for lauks in lauki:
            val = self.lauki_entries[lauks].get().strip()
            if lauks == "Datums":
                if not val:
                    messagebox.showwarning("Uzmanību", "Lūdzu ievadiet datumu!")
                    return
                try:
                    datetime.strptime(val, "%d.%m.%Y")
                except ValueError:
                    messagebox.showerror("Kļūda",
                        "Nepareizs datuma formāts!\n"
                        "Jāraksta: DD.MM.YYYY\n"
                        "Piemērs: 15.03.2025")
                    return
                vertibas.append(val)
            elif lauks == "Cena":
                if val == "":
                    vertibas.append(None)
                else:
                    try:
                        vertibas.append(float(val.replace(",", ".")))
                    except:
                        messagebox.showerror("Kļūda",
                            f"Lauks 'Cena' — jāievada skaitlis!\n"
                            f"Piemērs: 3.50 vai 3,50")
                        return
            else:
                vertibas.append(val if val else None)

        self.poga.config(state="disabled", text="Saglabā...")
        self.root.update()

        try:
            ok = saglabat_un_kartot(kategorija, vertibas)
            if ok:
                self.status_var.set("✓ Ieraksts pievienots un sakārtots!")
                self.status_label.config(fg="#27ae60")
                for lauks, entry in self.lauki_entries.items():
                    if lauks != "Datums":
                        entry.delete(0, tk.END)
        except Exception as e:
            messagebox.showerror("Kļūda", f"Neizdevās saglabāt:\n{e}")
            self.status_var.set("✗ Kļūda saglabājot!")
            self.status_label.config(fg="#e74c3c")

        self.poga.config(state="normal", text="Pievienot ierakstu")


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelPaligs(root)
    root.mainloop()
